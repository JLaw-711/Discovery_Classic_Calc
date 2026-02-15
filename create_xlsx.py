import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

SRC = Path("anaesthetic-billing-calculator-v2.txt")
OUT = Path("anaesthetic_billing_2026.xlsx")

text = SRC.read_text(encoding="utf-8")

# Parse constants
rcf_an = float(re.search(r"const RCF_AN = ([0-9.]+);", text).group(1))
rcf_cl = float(re.search(r"const RCF_CL = ([0-9.]+);", text).group(1))
vat = float(re.search(r"const VAT = ([0-9.]+);", text).group(1))

# Parse PLANS
plans = []
for m in re.finditer(r"\{\s*id:\"([^\"]+)\",\s*label:\"([^\"]+)\",\s*m:([0-9.]+),\s*loc:\"([^\"]+)\"\s*\}", text):
    plans.append((m.group(1), m.group(2), float(m.group(3)), m.group(4)))

# Parse PROCS entries [code,desc,r,a,s]
procs = []
for m in re.finditer(r"\[\"([0-9A-Za-z]+)\",\"([^\"]+)\",\s*([0-9.]+),\s*([0-9.]+),\s*(\d+)\]", text):
    procs.append((m.group(1), m.group(2), float(m.group(3)), float(m.group(4)), int(m.group(5))))

# Parse MODS
mods = []
for m in re.finditer(r"\{\s*c:\"([^\"]+)\",\s*d:\"([^\"]+)\",\s*u:([0-9.]+),\s*t:\"([^\"]+)\",\s*cat:\"([^\"]+)\"(?:,note:\"([^\"]+)\")?(?:,tm:([0-9.]+))?\s*\}", text):
    mods.append((m.group(1), m.group(2), float(m.group(3)), m.group(4), m.group(5), m.group(6) or "", m.group(7) or ""))

# Parse CONSULTS
consults = []
for m in re.finditer(r"\{\s*c:\"([^\"]+)\",\s*d:\"([^\"]+)\",\s*ih:([0-9.]+),\s*oh:([0-9.]+)(?:,on:true)?\s*\}", text):
    consults.append((m.group(1), m.group(2), float(m.group(3)), float(m.group(4))))

wb = Workbook()
rates = wb.active
rates.title = "Rates"
input_ws = wb.create_sheet("Input")
calc = wb.create_sheet("Calculations")
out = wb.create_sheet("Output")

# Styles
title_font = Font(size=14, bold=True, color="FFFFFF")
title_fill = PatternFill("solid", fgColor="1F4E78")
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2F5597")
label_fill = PatternFill("solid", fgColor="F2F2F2")
workings_fill = PatternFill("solid", fgColor="D9E1F2")
thin = Side(border_style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Config constants
rates["A1"] = "RCF_AN"
rates["B1"] = rcf_an
rates["A2"] = "RCF_CL"
rates["B2"] = rcf_cl
rates["A3"] = "VAT"
rates["B3"] = vat
rates["A4"] = "BASE_AN"
rates["B4"] = rcf_an / 2.04
rates["A5"] = "BASE_CL"
rates["B5"] = rcf_cl / 2.04

# Named constants
wb.defined_names.add(DefinedName("RCF_AN", attr_text="Rates!$B$1"))
wb.defined_names.add(DefinedName("RCF_CL", attr_text="Rates!$B$2"))
wb.defined_names.add(DefinedName("VAT", attr_text="Rates!$B$3"))
wb.defined_names.add(DefinedName("BASE_AN", attr_text="Rates!$B$4"))
wb.defined_names.add(DefinedName("BASE_CL", attr_text="Rates!$B$5"))

# Write PROCS table
row = 7
rates[f"A{row}"] = "Code"
rates[f"B{row}"] = "Description"
rates[f"C{row}"] = "RVU"
rates[f"D{row}"] = "BaseAmount"
rates[f"E{row}"] = "SpecialtyIndex"
start_procs_row = row + 1
for i, p in enumerate(procs, start=start_procs_row):
    rates[f"A{i}"] = p[0]
    rates[f"B{i}"] = p[1]
    rates[f"C{i}"] = p[2]
    rates[f"D{i}"] = p[3]
    rates[f"E{i}"] = p[4]
end_procs_row = start_procs_row + len(procs) - 1
rates.add_table(Table(displayName="tbl_PROCS", ref=f"A{row}:E{end_procs_row}"))

# PLANS
prow = end_procs_row + 2
rates[f"A{prow}"] = "PlanID"
rates[f"B{prow}"] = "Label"
rates[f"C{prow}"] = "Multiplier"
rates[f"D{prow}"] = "Loc"
for i, p in enumerate(plans, start=prow + 1):
    rates[f"A{i}"] = p[0]
    rates[f"B{i}"] = p[1]
    rates[f"C{i}"] = p[2]
    rates[f"D{i}"] = p[3]
end_plans_row = prow + len(plans)
rates.add_table(Table(displayName="tbl_PLANS", ref=f"A{prow}:D{end_plans_row}"))

# MODS
mrow = end_plans_row + 2
rates[f"A{mrow}"] = "ModCode"
rates[f"B{mrow}"] = "Description"
rates[f"C{mrow}"] = "Units"
rates[f"D{mrow}"] = "UnitType"
rates[f"E{mrow}"] = "Category"
for i, m in enumerate(mods, start=mrow + 1):
    rates[f"A{i}"] = m[0]
    rates[f"B{i}"] = m[1]
    rates[f"C{i}"] = m[2]
    rates[f"D{i}"] = m[3]
    rates[f"E{i}"] = m[4]
end_mods_row = mrow + len(mods)
rates.add_table(Table(displayName="tbl_MODS", ref=f"A{mrow}:E{end_mods_row}"))

# CONSULTS
crow = end_mods_row + 2
rates[f"A{crow}"] = "ConsCode"
rates[f"B{crow}"] = "Description"
rates[f"C{crow}"] = "IH"
rates[f"D{crow}"] = "OH"
for i, c in enumerate(consults, start=crow + 1):
    rates[f"A{i}"] = c[0]
    rates[f"B{i}"] = c[1]
    rates[f"C{i}"] = c[2]
    rates[f"D{i}"] = c[3]
end_cons_row = crow + len(consults)
rates.add_table(Table(displayName="tbl_CONSULTS", ref=f"A{crow}:D{end_cons_row}"))

# Rates column widths
for col in range(1, 6):
    rates.column_dimensions[get_column_letter(col)].width = 30

# INPUT sheet layout
input_ws["A1"] = "Patient"
input_ws["B1"] = ""
input_ws["A2"] = "ICD-10"
input_ws["B2"] = ""
input_ws["A3"] = "Surgeon"
input_ws["B3"] = ""

input_ws["A5"] = "Rate Plan"
input_ws["B5"] = ""

input_ws["A7"] = "Procedure Code"
input_ws["B7"] = ""
input_ws["C7"] = "=IF(B7=\"\",\"\",VLOOKUP(B7, Rates!$A${}:$B${}, 2, FALSE))".format(start_procs_row, end_procs_row)

input_ws["A9"] = "Anaesthetic Minutes"
input_ws["B9"] = ""

input_ws["A11"] = "Emergency? (Yes/No)"
input_ws["B11"] = "No"
input_ws["C11"] = "Emergency Minutes"
input_ws["D11"] = ""

input_ws["A13"] = "Consult 1"
input_ws["B13"] = ""
input_ws["A14"] = "Consult 2"
input_ws["B14"] = ""
input_ws["A15"] = "Consult 3"
input_ws["B15"] = ""

input_ws["A17"] = "Modifier Code"
input_ws["B17"] = "Enabled (Yes/No)"
for i in range(18, 38):
    input_ws[f"A{i}"] = ""
    input_ws[f"B{i}"] = "No"

# Data validations
plans_start = prow + 1
plans_end = end_plans_row
plan_range = f"'{rates.title}'!$B${plans_start}:$B${plans_end}"
dv_plan = DataValidation(type="list", formula1=plan_range, allow_blank=False)
input_ws.add_data_validation(dv_plan)
dv_plan.add(input_ws["B5"])

proc_start = start_procs_row
proc_end = end_procs_row
proc_range = f"'{rates.title}'!$A${proc_start}:$A${proc_end}"
dv_proc = DataValidation(type="list", formula1=proc_range, allow_blank=True)
input_ws.add_data_validation(dv_proc)
dv_proc.add(input_ws["B7"])

cons_start = crow + 1
cons_end = end_cons_row
cons_range = f"'{rates.title}'!$A${cons_start}:$A${cons_end}"
dv_cons = DataValidation(type="list", formula1=cons_range, allow_blank=True)
input_ws.add_data_validation(dv_cons)
dv_cons.add(input_ws["B13"])
dv_cons.add(input_ws["B14"])
dv_cons.add(input_ws["B15"])

mod_start = mrow + 1
mod_end = end_mods_row
modcode_range = f"'{rates.title}'!$A${mod_start}:$A${mod_end}"
dv_mod = DataValidation(type="list", formula1=modcode_range, allow_blank=True)
input_ws.add_data_validation(dv_mod)
for i in range(18, 38):
    dv_mod.add(input_ws[f"A{i}"])

v_yn = '"Yes,No"'
dv_yn = DataValidation(type="list", formula1=v_yn, allow_blank=False)
input_ws.add_data_validation(dv_yn)
dv_yn.add(input_ws["B11"])
for i in range(18, 38):
    dv_yn.add(input_ws[f"B{i}"])

# CALCULATIONS sheet formulas
c = calc
c["A1"] = "Plan Multiplier"
c["B1"] = "=IF(Input!B5=\"\",\"\",INDEX(Rates!$C${}:$C${},MATCH(Input!B5,Rates!$B${}:$B${},0)))".format(prow + 1, end_plans_row, prow + 1, end_plans_row)
c["A2"] = "Plan Location"
c["B2"] = "=IF(Input!B5=\"\",\"\",INDEX(Rates!$D${}:$D${},MATCH(Input!B5,Rates!$B${}:$B${},0)))".format(prow + 1, end_plans_row, prow + 1, end_plans_row)

c["A3"] = "Proc RVU"
c["B3"] = "=IF(Input!B7=\"\",\"\",VLOOKUP(Input!B7, Rates!$A${}:$C${},3,FALSE))".format(start_procs_row, end_procs_row)
c["A4"] = "Proc UnitPrice_exclVAT"
c["B4"] = "=IF(B3=\"\",\"\",(B3*BASE_AN*B1)/(1+VAT))"

c["A6"] = "Anaes Minutes"
c["B6"] = "=IF(Input!B9=\"\",0,Input!B9)"
c["A7"] = "Has BMI"
c["B7"] = "=IF(COUNTIFS(Input!$A$18:$A$37,\"18\",Input!$B$18:$B$37,\"Yes\")>0,1,0)"
c["A8"] = "Effective Minutes"
c["B8"] = "=IF(B7=1,ROUND(B6*1.5,0),B6)"
c["A9"] = "Time Units"
c["B9"] = "=IF(B8<=0,0,IF(B8<=60,CEILING(B8/15,1)*2,8+CEILING((B8-60)/15,1)*3))"

c["A10"] = "Anaes Unit Price exclVAT"
c["B10"] = "=IF(B1=\"\",\"\",(BASE_AN*B1)/(1+VAT))"
c["A11"] = "Anaes Amount exclVAT"
c["B11"] = "=B9*B10"

c["A13"] = "Emergency Minutes"
c["B13"] = "=IF(Input!B11=\"Yes\",Input!D11,0)"
c["A14"] = "Emergency Blocks"
c["B14"] = "=IF(B13<=0,0,CEILING(B13/30,1))"
c["A15"] = "Clinical unit price exclVAT"
c["B15"] = "=(BASE_CL*B1)/(1+VAT)"
c["A16"] = "Emergency Unit Price per block exclVAT"
c["B16"] = "=B15*12"
c["A17"] = "Emergency Amount exclVAT"
c["B17"] = "=B14*B16"

mod_codes_range = f"Rates!$A${mrow+1}:$A${end_mods_row}"
mod_units_range = f"Rates!$C${mrow+1}:$C${end_mods_row}"
mod_type_range = f"Rates!$D${mrow+1}:$D${end_mods_row}"
c["A20"] = "Modifiers Amount exclVAT"
mod_formula = ('=SUMPRODUCT( (COUNTIF(Input!$A$18:$A$37, {mc})>0) * ( {mu} * ((({mt}="an")*BASE_AN) + (({mt}="cl")*BASE_CL)) * B1 / (1+VAT) ) )')
mod_formula = mod_formula.replace("{mc}", mod_codes_range).replace("{mu}", mod_units_range).replace("{mt}", mod_type_range)
c["B20"] = mod_formula

# OUTPUT sheet: formatted quote with workings panel
o = out
o["A1"] = "Anaesthetic Quote (2026)"
o.merge_cells("A1:F1")
o["A2"] = "Patient"
o["B2"] = "=Input!B1"
o["C2"] = "ICD-10"
o["D2"] = "=Input!B2"
o["E2"] = "Plan"
o["F2"] = "=Input!B5"
o["A3"] = "Surgeon"
o["B3"] = "=Input!B3"
o["C3"] = "Procedure"
o["D3"] = "=Input!B7"
o["E3"] = "Minutes"
o["F3"] = "=Input!B9"

o["A5"] = "Code"
o["B5"] = "Description"
o["C5"] = "Qty"
o["D5"] = "Unit Price exclVAT"
o["E5"] = "VAT%"
o["F5"] = "Amount exclVAT"

consult_desc = f"=IF(Input!{{cell}}=\"\",\"\",VLOOKUP(Input!{{cell}}, Rates!$A${crow}:$B${end_cons_row},2,FALSE))"
consult_price = f"=IF(Input!{{cell}}=\"\",0,IF(Calculations!B2=\"IH\",VLOOKUP(Input!{{cell}}, Rates!$A${crow}:$D${end_cons_row},3,FALSE),VLOOKUP(Input!{{cell}}, Rates!$A${crow}:$D${end_cons_row},4,FALSE))/(1+VAT))"

# Consult 1
o["A6"] = "=IF(Input!B13=\"\",\"\",Input!B13)"
o["B6"] = consult_desc.replace("{cell}", "B13")
o["C6"] = "=IF(Input!B13=\"\",0,1)"
o["D6"] = consult_price.replace("{cell}", "B13")
o["E6"] = "15%"
o["F6"] = "=C6*D6"

# Consult 2
o["A7"] = "=IF(Input!B14=\"\",\"\",Input!B14)"
o["B7"] = consult_desc.replace("{cell}", "B14")
o["C7"] = "=IF(Input!B14=\"\",0,1)"
o["D7"] = consult_price.replace("{cell}", "B14")
o["E7"] = "15%"
o["F7"] = "=C7*D7"

# Consult 3
o["A8"] = "=IF(Input!B15=\"\",\"\",Input!B15)"
o["B8"] = consult_desc.replace("{cell}", "B15")
o["C8"] = "=IF(Input!B15=\"\",0,1)"
o["D8"] = consult_price.replace("{cell}", "B15")
o["E8"] = "15%"
o["F8"] = "=C8*D8"

# Procedure
o["A9"] = "=IF(Input!B7=\"\",\"\",Input!B7)"
o["B9"] = "=IF(Input!B7=\"\",\"\",VLOOKUP(Input!B7, Rates!$A${}:$B${},2,FALSE))".format(start_procs_row, end_procs_row)
o["C9"] = "=IF(Input!B7=\"\",0,1)"
o["D9"] = "=IF(Calculations!B4=\"\",0,Calculations!B4)"
o["E9"] = "15%"
o["F9"] = "=C9*D9"

# Anaesthetic time 0023
o["A10"] = "0023"
o["B10"] = "=IF(Calculations!B9=0,\"\",CONCAT(\"Anaesthetic time \",Input!B9,\" min\"))"
o["C10"] = "=Calculations!B9"
o["D10"] = "=Calculations!B10"
o["E10"] = "15%"
o["F10"] = "=C10*D10"

# Emergency 0011
o["A11"] = "0011"
o["B11"] = "=IF(Calculations!B14=0,\"\",CONCAT(\"Emergency \",Input!D11,\" min\"))"
o["C11"] = "=Calculations!B14"
o["D11"] = "=Calculations!B16"
o["E11"] = "15%"
o["F11"] = "=C11*D11"

# Modifiers total
o["A12"] = "MODS"
o["B12"] = "Modifiers (total)"
o["C12"] = "=IF(Calculations!B20=0,0,1)"
o["D12"] = "=IF(Calculations!B20=0,0,Calculations!B20)"
o["E12"] = "15%"
o["F12"] = "=D12"

# Totals
o["E14"] = "Subtotal (excl VAT)"
o["F14"] = "=SUM(F6:F12)"
o["E15"] = "VAT (15%)"
o["F15"] = "=F14*VAT"
o["E16"] = "Total (incl VAT)"
o["F16"] = "=F14+F15"

# Workings panel
o["H1"] = "Workings"
o.merge_cells("H1:I1")
o["H2"] = "Plan Multiplier"
o["I2"] = "=Calculations!B1"
o["H3"] = "Plan Location"
o["I3"] = "=Calculations!B2"
o["H4"] = "Proc RVU"
o["I4"] = "=Calculations!B3"
o["H5"] = "Proc Unit Price (ex VAT)"
o["I5"] = "=Calculations!B4"
o["H6"] = "Anaes Minutes"
o["I6"] = "=Calculations!B6"
o["H7"] = "Effective Minutes"
o["I7"] = "=Calculations!B8"
o["H8"] = "Time Units"
o["I8"] = "=Calculations!B9"
o["H9"] = "Anaes Unit Price (ex VAT)"
o["I9"] = "=Calculations!B10"
o["H10"] = "Emergency Blocks"
o["I10"] = "=Calculations!B14"
o["H11"] = "Emergency Block Price (ex VAT)"
o["I11"] = "=Calculations!B16"
o["H12"] = "Modifiers Amount (ex VAT)"
o["I12"] = "=Calculations!B20"

# Formatting: Input sheet
input_ws.column_dimensions["A"].width = 26
input_ws.column_dimensions["B"].width = 22
input_ws.column_dimensions["C"].width = 42
input_ws.column_dimensions["D"].width = 20
input_ws.freeze_panes = "A5"
input_ws.sheet_view.showGridLines = False

label_cells = ["A1", "A2", "A3", "A5", "A7", "A9", "A11", "C11", "A13", "A14", "A15", "A17", "B17"]
for cell in label_cells:
    input_ws[cell].font = Font(bold=True)
    input_ws[cell].fill = label_fill
    input_ws[cell].alignment = Alignment(horizontal="left", vertical="center")
    input_ws[cell].border = border

input_cells = ["B1", "B2", "B3", "B5", "B7", "B9", "B11", "D11", "B13", "B14", "B15"]
for cell in input_cells:
    input_ws[cell].border = border
    input_ws[cell].alignment = Alignment(horizontal="left", vertical="center")

for i in range(18, 38):
    input_ws[f"A{i}"].border = border
    input_ws[f"B{i}"].border = border
    input_ws[f"A{i}"].alignment = Alignment(horizontal="left", vertical="center")
    input_ws[f"B{i}"].alignment = Alignment(horizontal="center", vertical="center")

input_ws["C7"].border = border
input_ws["C7"].alignment = Alignment(horizontal="left", vertical="center")

# Formatting: Output sheet
out.column_dimensions["A"].width = 12
out.column_dimensions["B"].width = 50
out.column_dimensions["C"].width = 8
out.column_dimensions["D"].width = 18
out.column_dimensions["E"].width = 8
out.column_dimensions["F"].width = 18
out.column_dimensions["H"].width = 28
out.column_dimensions["I"].width = 20
out.freeze_panes = "A6"
out.sheet_view.showGridLines = False

o["A1"].font = title_font
o["A1"].fill = title_fill
o["A1"].alignment = Alignment(horizontal="center", vertical="center")

for cell in ["A2", "C2", "E2", "A3", "C3", "E3"]:
    o[cell].font = Font(bold=True)

for col in range(1, 7):
    cell = out.cell(row=5, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

for row in range(6, 13):
    for col in range(1, 7):
        cell = out.cell(row=row, column=col)
        cell.border = border
        if col in (1, 3, 5):
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

for row in range(14, 17):
    for col in range(5, 7):
        cell = out.cell(row=row, column=col)
        cell.border = border
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if col == 5:
            cell.font = Font(bold=True)

currency_format = "R #,##0.00"
for row in range(6, 17):
    out.cell(row=row, column=4).number_format = currency_format
    out.cell(row=row, column=6).number_format = currency_format

# Workings formatting
o["H1"].font = header_font
o["H1"].fill = header_fill
o["H1"].alignment = Alignment(horizontal="center", vertical="center")
for row in range(2, 13):
    o[f"H{row}"].fill = workings_fill
    o[f"H{row}"].font = Font(bold=True)
    o[f"H{row}"].border = border
    o[f"I{row}"].border = border
    o[f"H{row}"].alignment = Alignment(horizontal="left", vertical="center")
    o[f"I{row}"].alignment = Alignment(horizontal="right", vertical="center")

for row in range(5, 13):
    out.cell(row=row, column=5).alignment = Alignment(horizontal="center", vertical="center")

# Hide calculations sheet for non-technical users
calc.sheet_state = "hidden"

# Save workbook
wb.save(OUT)
print(f"Written {OUT}")
